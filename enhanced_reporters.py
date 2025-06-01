"""Enhanced report generation functionality - Phase 1"""

import json
import pandas as pd
from datetime import datetime
from typing import Dict, List
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

from models import (
    Customer, PaymentPlan, CustomerIssue, PaymentMetrics,
    DataQualityReport, ErrorType, IssueSeverity
)

class EnhancedReportGenerator:
    """Enhanced report generator with error highlighting and class filtering"""
    
    def __init__(self, output_dir: str = './reports'):
        self.output_dir = output_dir
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Create output directory if it doesn't exist
        os.makedirs(self.output_dir, exist_ok=True)
        
    def generate_comprehensive_quality_report(self, 
                                           customers: Dict[str, Customer],
                                           clean_customers: List[Customer],
                                           problematic_customers: List[Customer],
                                           all_issues: List[CustomerIssue],
                                           data_quality_report: DataQualityReport) -> Dict:
        """Generate enhanced quality report"""
        
        # Count issues by type and severity
        issue_counts = {}
        severity_counts = {'critical': 0, 'warning': 0, 'info': 0}
        
        for issue in all_issues:
            issue_type = issue.issue_type.value
            issue_counts[issue_type] = issue_counts.get(issue_type, 0) + 1
            severity_counts[issue.severity.value] += 1
        
        # Calculate financial impact
        total_outstanding = sum(c.total_open_balance for c in customers.values())
        clean_outstanding = sum(c.total_open_balance for c in clean_customers)
        problematic_outstanding = sum(c.total_open_balance for c in problematic_customers)
        
        # Calculate impact by class
        class_impact = {}
        for customer in customers.values():
            for class_name in customer.all_classes:
                if class_name not in class_impact:
                    class_impact[class_name] = {
                        'total_customers': 0,
                        'clean_customers': 0,
                        'problematic_customers': 0,
                        'total_outstanding': 0
                    }
                
                class_impact[class_name]['total_customers'] += 1
                class_impact[class_name]['total_outstanding'] += customer.total_open_balance
                
                if customer in clean_customers:
                    class_impact[class_name]['clean_customers'] += 1
                else:
                    class_impact[class_name]['problematic_customers'] += 1
        
        report = {
            'summary': {
                'report_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'total_customers': len(customers),
                'clean_customers': len(clean_customers),
                'problematic_customers': len(problematic_customers),
                'total_payment_plans': sum(len(c.payment_plans) for c in customers.values()),
                'customers_with_multiple_plans': sum(1 for c in customers.values() if c.has_multiple_plans),
                'total_issues': len(all_issues),
                'total_outstanding': round(total_outstanding, 2),
                'clean_outstanding': round(clean_outstanding, 2),
                'problematic_outstanding': round(problematic_outstanding, 2),
                'percentage_with_issues': round((len(problematic_customers) / len(customers) * 100), 1) if customers else 0,
                'data_quality_score': round(((len(clean_customers) / len(customers)) * 100), 1) if customers else 0
            },
            'data_processing': {
                'total_rows_processed': data_quality_report.total_rows_processed,
                'total_invoices_processed': data_quality_report.total_invoices_processed,
                'invoices_with_open_balance': data_quality_report.total_invoices_with_open_balance,
                'invoices_ignored_zero_balance': data_quality_report.total_invoices_ignored,
                'classes_found': data_quality_report.classes_found
            },
            'issue_breakdown': issue_counts,
            'issue_severity_breakdown': severity_counts,
            'class_breakdown': class_impact,
            'recommendations': self._generate_enhanced_recommendations(issue_counts, problematic_customers, class_impact),
            'top_problematic_customers': self._get_top_problematic_customers(problematic_customers, 15),
            'critical_issues_requiring_immediate_attention': self._get_critical_issues_summary(all_issues)
        }
        
        return report
    
    def generate_enhanced_dashboard_data(self,
                                       customers: Dict[str, Customer],
                                       clean_customers: List[Customer],
                                       problematic_customers: List[Customer],
                                       all_metrics: List[PaymentMetrics]) -> Dict:
        """Generate enhanced dashboard data with multi-plan support"""
        
        # Calculate summary metrics
        total_outstanding_tracked = sum(m.total_owed for m in all_metrics)
        total_outstanding_untracked = sum(c.total_open_balance for c in problematic_customers)
        
        # Group metrics by customer for customer-level analysis
        customer_metrics = {}
        for metric in all_metrics:
            if metric.customer_name not in customer_metrics:
                customer_metrics[metric.customer_name] = []
            customer_metrics[metric.customer_name].append(metric)
        
        # Calculate expected monthly (normalized)
        expected_monthly = 0
        customers_behind = 0
        customers_current = 0
        customers_completed = 0
        
        customer_status_map = {}
        for customer_name, metrics_list in customer_metrics.items():
            # Determine overall customer status (worst case across all plans)
            customer_status = 'current'
            customer_expected_monthly = 0
            
            for metric in metrics_list:
                # Normalize to monthly
                if metric.frequency == 'monthly':
                    customer_expected_monthly += metric.monthly_payment
                elif metric.frequency == 'quarterly':
                    customer_expected_monthly += metric.monthly_payment / 3
                elif metric.frequency == 'bimonthly':
                    customer_expected_monthly += metric.monthly_payment / 2
                
                # Worst status wins
                if metric.status.value == 'behind':
                    customer_status = 'behind'
                elif metric.status.value == 'completed' and customer_status == 'current':
                    customer_status = 'completed'
            
            expected_monthly += customer_expected_monthly
            customer_status_map[customer_name] = customer_status
            
            if customer_status == 'behind':
                customers_behind += 1
            elif customer_status == 'completed':
                customers_completed += 1
            else:
                customers_current += 1
        
        dashboard_data = {
            'summary_metrics': {
                'total_customers_shown': len(customer_metrics),
                'total_customers_skipped': len(problematic_customers),
                'total_payment_plans_tracked': len(all_metrics),
                'total_outstanding_tracked': round(total_outstanding_tracked, 2),
                'total_outstanding_untracked': round(total_outstanding_untracked, 2),
                'expected_monthly_collection': round(expected_monthly, 2),
                'customers_behind': customers_behind,
                'customers_current': customers_current,
                'customers_completed': customers_completed,
                'percentage_behind': round((customers_behind / len(customer_metrics) * 100), 1) if customer_metrics else 0
            },
            'customer_summaries': self._generate_customer_summaries(customer_metrics),
            'payment_plan_details': [self._metrics_to_dict(m) for m in all_metrics],
            'skipped_customers': [self._problem_customer_summary(c) for c in problematic_customers],
            'class_summaries': self._generate_class_summaries(all_metrics),
            'payment_roadmaps': self._generate_roadmap_summaries(all_metrics)
        }
        
        # Sort data
        dashboard_data['customer_summaries'].sort(key=lambda x: x.get('worst_months_behind', 0), reverse=True)
        dashboard_data['payment_plan_details'].sort(key=lambda x: x.get('months_behind', 0), reverse=True)
        dashboard_data['skipped_customers'].sort(key=lambda x: x['total_open'], reverse=True)
        
        return dashboard_data
    
    def generate_error_highlighted_excel(self, 
                                       customers: Dict[str, Customer],
                                       all_issues: List[CustomerIssue],
                                       filename: str = None) -> str:
        """Generate Excel file with error highlighting"""
        
        if not filename:
            filename = f'payment_plan_errors_{self.timestamp}.xlsx'
        
        filepath = os.path.join(self.output_dir, filename)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default worksheet
        wb.remove(wb.active)
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Error Summary")
        self._create_error_summary_sheet(summary_ws, all_issues)
        
        # Create detailed data sheet with highlighting
        data_ws = wb.create_sheet("Data with Errors Highlighted")
        self._create_highlighted_data_sheet(data_ws, customers, all_issues)
        
        # Create issues by customer sheet
        issues_ws = wb.create_sheet("Issues by Customer")
        self._create_issues_by_customer_sheet(issues_ws, all_issues)
        
        # Create class analysis sheet
        class_ws = wb.create_sheet("Analysis by Class")
        self._create_class_analysis_sheet(class_ws, customers)
        
        # Save workbook
        wb.save(filepath)
        return filepath
    
    def _create_error_summary_sheet(self, ws, all_issues: List[CustomerIssue]):
        """Create error summary sheet"""
        # Headers
        headers = ['Error Type', 'Severity', 'Count', 'Affected Customers', 'Common Fix']
        ws.append(headers)
        
        # Make headers bold
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Group issues
        issue_summary = {}
        for issue in all_issues:
            key = (issue.issue_type.value, issue.severity.value)
            if key not in issue_summary:
                issue_summary[key] = {
                    'count': 0,
                    'customers': set(),
                    'suggested_fix': issue.suggested_fix
                }
            issue_summary[key]['count'] += 1
            issue_summary[key]['customers'].add(issue.customer_name)
        
        # Add data
        for (issue_type, severity), data in issue_summary.items():
            row = [
                issue_type.replace('_', ' ').title(),
                severity.upper(),
                data['count'],
                len(data['customers']),
                data['suggested_fix'] or 'Review manually'
            ]
            ws.append(row)
            
            # Color code by severity
            row_num = ws.max_row
            if severity == 'critical':
                fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            elif severity == 'warning':
                fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            else:
                fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            
            for cell in ws[row_num]:
                cell.fill = fill
    
    def _create_highlighted_data_sheet(self, ws, customers: Dict[str, Customer], all_issues: List[CustomerIssue]):
        """Create data sheet with error highlighting"""
        # Headers
        headers = ['Customer Name', 'Plan ID', 'Invoice Number', 'Date', 'Payment Terms', 
                  'Original Amount', 'Open Balance', 'Class', 'Issues Found']
        ws.append(headers)
        
        # Make headers bold
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Create issue lookup for quick access
        issues_by_customer = {}
        for issue in all_issues:
            if issue.customer_name not in issues_by_customer:
                issues_by_customer[issue.customer_name] = []
            issues_by_customer[issue.customer_name].append(issue)
        
        # Add data with highlighting
        for customer_name, customer in customers.items():
            customer_issues = issues_by_customer.get(customer_name, [])
            
            for plan in customer.payment_plans:
                for invoice in plan.invoices:
                    # Find issues affecting this invoice
                    invoice_issues = []
                    for issue in customer_issues:
                        if invoice.invoice_number in issue.affected_invoices or not issue.affected_invoices:
                            invoice_issues.append(f"{issue.issue_type.value}: {issue.description}")
                    
                    row_data = [
                        customer_name,
                        plan.plan_id,
                        invoice.invoice_number,
                        invoice.date.strftime('%Y-%m-%d') if invoice.date else '',
                        invoice.payment_terms or '',
                        invoice.original_amount,
                        invoice.open_balance,
                        invoice.class_field or '',
                        '; '.join(invoice_issues) if invoice_issues else ''
                    ]
                    
                    ws.append(row_data)
                    
                    # Highlight row if there are issues
                    if invoice_issues:
                        row_num = ws.max_row
                        # Determine highlight color based on severity
                        has_critical = any('critical' in str(issue) for issue in customer_issues)
                        has_warning = any('warning' in str(issue) for issue in customer_issues)
                        
                        if has_critical:
                            fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                        elif has_warning:
                            fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                        else:
                            fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                        
                        for cell in ws[row_num]:
                            cell.fill = fill
    
    def _create_issues_by_customer_sheet(self, ws, all_issues: List[CustomerIssue]):
        """Create issues by customer sheet"""
        headers = ['Customer', 'Issue Type', 'Severity', 'Description', 'Suggested Fix', 'Field', 'Current Value']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Sort issues by customer, then severity
        sorted_issues = sorted(all_issues, key=lambda x: (x.customer_name, x.severity.value))
        
        for issue in sorted_issues:
            row = [
                issue.customer_name,
                issue.issue_type.value.replace('_', ' ').title(),
                issue.severity.value.upper(),
                issue.description,
                issue.suggested_fix or '',
                issue.field_name or '',
                issue.current_value or ''
            ]
            ws.append(row)
            
            # Color code by severity
            row_num = ws.max_row
            if issue.severity == IssueSeverity.CRITICAL:
                fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            elif issue.severity == IssueSeverity.WARNING:
                fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            else:
                fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
            
            for cell in ws[row_num]:
                cell.fill = fill
    
    def _create_class_analysis_sheet(self, ws, customers: Dict[str, Customer]):
        """Create class analysis sheet"""
        headers = ['Class', 'Total Customers', 'Total Plans', 'Total Outstanding', 'Avg per Customer']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        # Analyze by class
        class_analysis = {}
        for customer in customers.values():
            for class_name in customer.all_classes:
                if class_name not in class_analysis:
                    class_analysis[class_name] = {
                        'customers': set(),
                        'total_plans': 0,
                        'total_outstanding': 0
                    }
                
                class_analysis[class_name]['customers'].add(customer.customer_name)
                class_analysis[class_name]['total_plans'] += len(customer.payment_plans)
                class_analysis[class_name]['total_outstanding'] += customer.total_open_balance
        
        # Add data
        for class_name, data in sorted(class_analysis.items()):
            customer_count = len(data['customers'])
            avg_per_customer = data['total_outstanding'] / customer_count if customer_count > 0 else 0
            
            row = [
                class_name,
                customer_count,
                data['total_plans'],
                round(data['total_outstanding'], 2),
                round(avg_per_customer, 2)
            ]
            ws.append(row)
    
    def _generate_customer_summaries(self, customer_metrics: Dict[str, List[PaymentMetrics]]) -> List[Dict]:
        """Generate customer-level summaries from multiple plans"""
        summaries = []
        
        for customer_name, metrics_list in customer_metrics.items():
            total_owed = sum(m.total_owed for m in metrics_list)
            total_original = sum(m.original_amount for m in metrics_list)
            worst_months_behind = max((m.months_behind for m in metrics_list), default=0)
            
            # Determine overall status
            statuses = [m.status.value for m in metrics_list]
            if 'behind' in statuses:
                overall_status = 'behind'
            elif 'completed' in statuses and all(s in ['completed', 'current'] for s in statuses):
                overall_status = 'completed'
            else:
                overall_status = 'current'
            
            # Calculate total expected monthly
            total_expected_monthly = 0
            for m in metrics_list:
                if m.frequency == 'monthly':
                    total_expected_monthly += m.monthly_payment
                elif m.frequency == 'quarterly':
                    total_expected_monthly += m.monthly_payment / 3
                elif m.frequency == 'bimonthly':
                    total_expected_monthly += m.monthly_payment / 2
            
            summary = {
                'customer_name': customer_name,
                'total_plans': len(metrics_list),
                'total_owed': round(total_owed, 2),
                'total_original': round(total_original, 2),
                'percent_paid': round(((total_original - total_owed) / total_original * 100), 1) if total_original > 0 else 0,
                'worst_months_behind': round(worst_months_behind, 1),
                'overall_status': overall_status,
                'total_expected_monthly': round(total_expected_monthly, 2),
                'plan_details': [self._metrics_to_dict(m) for m in metrics_list]
            }
            
            summaries.append(summary)
        
        return summaries
    
    def _generate_class_summaries(self, all_metrics: List[PaymentMetrics]) -> Dict[str, Dict]:
        """Generate summaries by class"""
        class_summaries = {}
        
        for metric in all_metrics:
            class_key = metric.class_field or 'Unknown'
            
            if class_key not in class_summaries:
                class_summaries[class_key] = {
                    'total_plans': 0,
                    'total_owed': 0,
                    'customers_behind': 0,
                    'expected_monthly': 0,
                    'customers': set()
                }
            
            class_summaries[class_key]['total_plans'] += 1
            class_summaries[class_key]['total_owed'] += metric.total_owed
            class_summaries[class_key]['customers'].add(metric.customer_name)
            
            if metric.status.value == 'behind':
                class_summaries[class_key]['customers_behind'] += 1
            
            # Normalize to monthly
            if metric.frequency == 'monthly':
                class_summaries[class_key]['expected_monthly'] += metric.monthly_payment
            elif metric.frequency == 'quarterly':
                class_summaries[class_key]['expected_monthly'] += metric.monthly_payment / 3
            elif metric.frequency == 'bimonthly':
                class_summaries[class_key]['expected_monthly'] += metric.monthly_payment / 2
        
        # Convert to final format
        for class_key in class_summaries:
            class_summaries[class_key]['total_customers'] = len(class_summaries[class_key]['customers'])
            del class_summaries[class_key]['customers']  # Remove set
        
        return class_summaries
    
    def _generate_roadmap_summaries(self, all_metrics: List[PaymentMetrics]) -> Dict[str, List[Dict]]:
        """Generate roadmap summaries for customers"""
        roadmaps = {}
        
        for metric in all_metrics:
            if metric.customer_name not in roadmaps:
                roadmaps[metric.customer_name] = []
            
            if metric.payment_roadmap:
                # Take first 6 payments for summary
                roadmap_summary = {
                    'plan_id': metric.plan_id,
                    'next_payments': metric.payment_roadmap[:6],
                    'total_payments_remaining': len(metric.payment_roadmap)
                }
                roadmaps[metric.customer_name].append(roadmap_summary)
        
        return roadmaps
    
    def _metrics_to_dict(self, metrics: PaymentMetrics) -> Dict:
        """Convert metrics to dictionary for JSON"""
        return {
            'customer_name': metrics.customer_name,
            'plan_id': metrics.plan_id,
            'monthly_payment': round(metrics.monthly_payment, 2),
            'frequency': metrics.frequency,
            'total_owed': round(metrics.total_owed, 2),
            'original_amount': round(metrics.original_amount, 2),
            'percent_paid': round(metrics.percent_paid, 1),
            'months_elapsed': metrics.months_elapsed,
            'expected_payments': metrics.expected_payments,
            'actual_payments': metrics.actual_payments,
            'payment_difference': metrics.payment_difference,
            'months_behind': metrics.months_behind,
            'status': metrics.status.value,
            'projected_completion': metrics.projected_completion.strftime('%Y-%m-%d') if metrics.projected_completion else None,
            'months_remaining': metrics.months_remaining,
            'class_field': metrics.class_field,
            'payment_roadmap': metrics.payment_roadmap
        }
    
    def _problem_customer_summary(self, customer: Customer) -> Dict:
        """Create summary for problematic customer"""
        all_issues = []
        critical_issues = []
        
        for plan in customer.payment_plans:
            for issue in plan.issues:
                all_issues.append(issue.issue_type.value)
                if issue.severity == IssueSeverity.CRITICAL:
                    critical_issues.append(issue.issue_type.value)
        
        return {
            'customer_name': customer.customer_name,
            'total_open': round(customer.total_open_balance, 2),
            'total_plans': len(customer.payment_plans),
            'all_classes': customer.all_classes,
            'issues': all_issues,
            'critical_issues': critical_issues,
            'issue_descriptions': [issue.description for plan in customer.payment_plans for issue in plan.issues]
        }
    
    def _generate_enhanced_recommendations(self, issue_counts: Dict, problematic_customers: List[Customer], class_impact: Dict) -> List[Dict]:
        """Generate enhanced recommendations"""
        recommendations = []
        
        if issue_counts.get('no_payment_terms', 0) > 0:
            affected_balance = sum(c.total_open_balance for c in problematic_customers 
                                 if any(any(i.issue_type.value == 'no_payment_terms' for i in plan.issues) 
                                       for plan in c.payment_plans))
            recommendations.append({
                'priority': 1,
                'action': 'Add payment terms to customers without them',
                'impact': f'{issue_counts["no_payment_terms"]} payment plans cannot be properly tracked',
                'affected_balance': round(affected_balance, 2),
                'effort': 'Low - Add payment amount and frequency to FOB field',
                'urgency': 'High - Prevents payment tracking entirely'
            })
        
        if issue_counts.get('multiple_payment_terms', 0) > 0:
            recommendations.append({
                'priority': 2,
                'action': 'Standardize payment terms for customers with multiple payment plans',
                'impact': f'{issue_counts["multiple_payment_terms"]} customers have ambiguous payment schedules',
                'effort': 'Medium - Contact customers to confirm current arrangement',
                'urgency': 'Medium - May lead to collection confusion'
            })
        
        # Add class-specific recommendations
        if class_impact:
            worst_class = max(class_impact.items(), 
                            key=lambda x: x[1]['problematic_customers'] / x[1]['total_customers'] if x[1]['total_customers'] > 0 else 0)
            if worst_class[1]['problematic_customers'] > 0:
                recommendations.append({
                    'priority': 3,
                    'action': f'Focus data cleanup efforts on {worst_class[0]} class',
                    'impact': f'{worst_class[1]["problematic_customers"]} out of {worst_class[1]["total_customers"]} customers in this class have issues',
                    'effort': 'Medium - Targeted cleanup approach',
                    'urgency': 'Medium - Improves overall data quality'
                })
        
        return recommendations
    
    def _get_top_problematic_customers(self, problematic_customers: List[Customer], limit: int = 15) -> List[Dict]:
        """Get top problematic customers by balance"""
        sorted_customers = sorted(problematic_customers, key=lambda c: c.total_open_balance, reverse=True)[:limit]
        
        result = []
        for customer in sorted_customers:
            all_issues = []
            critical_issues = []
            
            for plan in customer.payment_plans:
                for issue in plan.issues:
                    all_issues.append(issue.issue_type.value)
                    if issue.severity == IssueSeverity.CRITICAL:
                        critical_issues.append(issue.issue_type.value)
            
            result.append({
                'customer_name': customer.customer_name,
                'total_open': round(customer.total_open_balance, 2),
                'total_plans': len(customer.payment_plans),
                'issue_count': len(all_issues),
                'critical_issues': list(set(critical_issues)),
                'all_issues': list(set(all_issues)),
                'classes': customer.all_classes
            })
        
        return result
    
    def _get_critical_issues_summary(self, all_issues: List[CustomerIssue]) -> List[Dict]:
        """Get summary of critical issues requiring immediate attention"""
        critical_issues = [issue for issue in all_issues if issue.severity == IssueSeverity.CRITICAL]
        
        # Group by type
        critical_summary = {}
        for issue in critical_issues:
            issue_type = issue.issue_type.value
            if issue_type not in critical_summary:
                critical_summary[issue_type] = {
                    'count': 0,
                    'customers_affected': set(),
                    'total_impact': 0,
                    'example_customer': None
                }
            
            critical_summary[issue_type]['count'] += 1
            critical_summary[issue_type]['customers_affected'].add(issue.customer_name)
            critical_summary[issue_type]['example_customer'] = issue.customer_name
        
        result = []
        for issue_type, data in critical_summary.items():
            result.append({
                'issue_type': issue_type.replace('_', ' ').title(),
                'count': data['count'],
                'customers_affected': len(data['customers_affected']),
                'example_customer': data['example_customer'],
                'urgency': 'Immediate - Prevents accurate payment tracking'
            })
        
        return sorted(result, key=lambda x: x['customers_affected'], reverse=True)
    
    def save_all_reports(self, 
                        quality_report: Dict,
                        dashboard_data: Dict,
                        all_issues: List[CustomerIssue],
                        all_metrics: List[PaymentMetrics],
                        customers: Dict[str, Customer]):
        """Save all enhanced reports"""
        
        # Save quality report JSON
        with open(os.path.join(self.output_dir, f'enhanced_quality_report_{self.timestamp}.json'), 'w') as f:
            json.dump(quality_report, f, indent=2, default=str)
        
        # Save dashboard data JSON
        with open(os.path.join(self.output_dir, f'enhanced_dashboard_data_{self.timestamp}.json'), 'w') as f:
            json.dump(dashboard_data, f, indent=2, default=str)
        
        # Save enhanced error Excel file
        error_excel_path = self.generate_error_highlighted_excel(customers, all_issues)
        
        # Save metrics CSV
        if all_metrics:
            metrics_df = pd.DataFrame([self._metrics_to_dict(m) for m in all_metrics])
            metrics_df.to_csv(
                os.path.join(self.output_dir, f'payment_plan_metrics_{self.timestamp}.csv'), 
                index=False
            )
        
        # Save issues CSV
        if all_issues:
            issues_df = pd.DataFrame([issue.to_dict() for issue in all_issues])
            issues_df.to_csv(
                os.path.join(self.output_dir, f'all_issues_{self.timestamp}.csv'), 
                index=False
            )
        
        print(f"\nEnhanced reports saved to {self.output_dir}/")
        print(f"Timestamp: {self.timestamp}")
        print(f"Error-highlighted Excel: {os.path.basename(error_excel_path)}")
        return self.timestamp